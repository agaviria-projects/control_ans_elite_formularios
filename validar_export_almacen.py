# ============================================================
# 1. LIBRERÍAS
# ============================================================
from pathlib import Path
import pandas as pd
import time
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ============================================================
# 2. CONFIGURACIÓN DE RUTAS
# ============================================================
base = Path(__file__).resolve().parent

# Detecta automáticamente si existe el archivo TXT o XLSX
ruta_fenix_txt = base / "data_raw" / "Digitacion Fenix.txt"
ruta_fenix_xlsx = base / "data_raw" / "Digitacion Fenix.xlsx"

# Usa el archivo que realmente exista
if ruta_fenix_txt.exists():
    ruta_fenix = ruta_fenix_txt
    print("📁 Detectado archivo Fénix: Digitacion Fenix.txt")
else:
    ruta_fenix = ruta_fenix_xlsx
    print("📁 Detectado archivo Fénix: Digitacion Fenix.xlsx")

ruta_elite = base / "data_raw" / "Planilla Consumos.xlsx"
ruta_salida = base / "data_clean" / "CONTROL_ALMACEN.xlsx"


print("------------------------------------------------------------")
print("🚀 INICIANDO CRUCE FÉNIX vs ELITE (v3.2)...")
print("------------------------------------------------------------")

# ============================================================
# 3. CARGA DE DATOS
# ============================================================

columnas_fenix = [
    "pedido", "subz", "municipio", "contrato", "acta", "actividad",
    "fecha_estado", "pagina", "urbrur", "tipre", "red_interna",
    "tipo_operacion", "tipo", "cobro", "suminis", "item_cont",
    "item_res", "cantidad", "vlr_cliente", "valor_costo"
]

# --- FÉNIX ---
try:
    if ruta_fenix.suffix.lower() == ".txt":
        try:
            # Intento 1: lectura estándar UTF-8
            df_fenix = pd.read_csv(ruta_fenix, sep=None, engine="python", dtype=str, encoding="utf-8")
        except UnicodeDecodeError:
            # Intento 2: lectura alternativa en Latin-1 (Windows)
            df_fenix = pd.read_csv(ruta_fenix, sep=None, engine="python", dtype=str, encoding="latin-1")
            print("⚙️ Archivo leído correctamente con codificación Latin-1.")
    else:
        df_fenix = pd.read_excel(ruta_fenix, dtype=str)

    # Normalización de columnas
    df_fenix.columns = df_fenix.columns.str.lower().str.strip()
    df_fenix = df_fenix[[c for c in columnas_fenix if c in df_fenix.columns]]
    df_fenix["cantidad_fenix"] = pd.to_numeric(df_fenix["cantidad"], errors="coerce").fillna(0)
    if "mano_obra" not in df_fenix.columns:
        df_fenix["mano_obra"] = None

except Exception as e:
    raise SystemExit(f"❌ Error al leer FÉNIX: {e}")

# --- ELITE ---
try:
    print("🔎 Buscando hoja con encabezado 'CANTIDAD' y limpiando filas previas...")

    temp = pd.read_excel(ruta_elite, sheet_name=None, dtype=str, header=None)
    hoja_correcta, fila_header = None, None

    hojas_excluir = ["pasos", "resumen", "hoja1", "macro", "configuracion"]

    for hoja, df_temp in temp.items():
        if hoja.lower() in hojas_excluir:
            continue

        # Buscar fila con la palabra 'cantidad' (encabezado real)
        for i, fila in df_temp.iterrows():
            fila_texto = " ".join(str(x).lower() for x in fila.values if pd.notna(x))
            if "cantidad" in fila_texto:
                hoja_correcta, fila_header = hoja, i
                break
        if hoja_correcta:
            break

    if hoja_correcta is None:
        raise Exception("No se encontró ninguna hoja con encabezado 'CANTIDAD'.")

    print(f"📍 Hoja detectada: {hoja_correcta} | Encabezado en fila {fila_header + 1}")

    # 🔹 Leer solo desde el encabezado hacia abajo (evita filas vacías previas)
    df_elite = pd.read_excel(
        ruta_elite,
        sheet_name=hoja_correcta,
        dtype=str,
        skiprows=fila_header  # <-- lee desde esa fila hacia abajo
    )

    # 🔹 Limpiar encabezados
    df_elite.columns = (
        df_elite.columns
        .map(str)
        .str.lower()
        .str.strip()
        .str.replace(r"unnamed.*", "", regex=True)
    )

    # 🔹 Quitar columnas completamente vacías
    df_elite = df_elite.dropna(axis=1, how="all")

    # 🔹 Conservar columnas relevantes
    posibles_cols = ["#pedido", "pedido", "codigo", "cantidad", "descripción", "descripcion", "unidad", "tecnico"]
    df_elite = df_elite[[c for c in df_elite.columns if any(p in c for p in posibles_cols)]]

    # 🔹 Renombrar columnas estándar
    df_elite.rename(columns={
        "#pedido": "pedido",
        "codigo": "codigo",
        "cantidad": "cantidad_elite"
    }, inplace=True)

    # 🔹 Filtrar filas vacías o corruptas (pedidos falsos)
    df_elite["pedido"] = df_elite["pedido"].astype(str).str.strip()
    df_elite = df_elite[
        df_elite["pedido"].notna() &
        df_elite["pedido"].str.match(r"^\d{8,}$", na=False)
    ]

    # 🔹 Limpiar códigos válidos (6 dígitos)
    df_elite = df_elite[df_elite["codigo"].astype(str).str.match(r"^\d{6}$", na=False)]

    # 🔹 Convertir cantidades a número
    df_elite["cantidad_elite"] = pd.to_numeric(df_elite["cantidad_elite"], errors="coerce").fillna(0)

    # 🔹 Agregar columna mano_obra si no existe
    if "mano_obra" not in df_elite.columns:
        df_elite["mano_obra"] = None

    print(f"✅ Planilla Consumos lista para cruce: {len(df_elite)} registros limpios.")

except Exception as e:
    raise SystemExit(f"❌ Error al leer Planilla Consumos: {e}")



print("✅ Archivos cargados correctamente.")
time.sleep(0.5)

# ============================================================
# 4. CRUCE PRINCIPAL (FÉNIX vs ELITE)
# ============================================================
print("⚙️ Ejecutando cruce principal FÉNIX vs ELITE...")

df_fenix.rename(columns={"item_res": "codigo"}, inplace=True)

# Validar columnas clave
for col in ["pedido", "codigo"]:
    if col not in df_fenix.columns:
        df_fenix[col] = None
    if col not in df_elite.columns:
        df_elite[col] = None

# Filtrar códigos válidos (solo 6 dígitos)
df_elite = df_elite[df_elite["codigo"].astype(str).str.match(r"^\d{6}$", na=False)]

# ============================================================
# 4.1. Normalizar códigos base y complementarios antes del merge
# ============================================================

# 🔹 Definir equivalencias base ↔ complemento
equivalencias = {
    "200492A": "200492",
    "200384A": "200384"
}

# 🔹 Crear columna auxiliar con el código base normalizado
df_fenix["codigo_equiv"] = df_fenix["codigo"].replace(equivalencias)
df_elite["codigo_equiv"] = df_elite["codigo"].replace(equivalencias)

# 🔹 Agregar columna 'origen' antes del merge (evita KeyError)
df_fenix["origen"] = "FENIX"
df_elite["origen"] = "ELITE"

# 🔹 Merge extendido usando el código normalizado
df_full = pd.merge(
    df_fenix,
    df_elite[["pedido", "codigo_equiv", "cantidad_elite", "origen"]],
    left_on=["pedido", "codigo_equiv"],
    right_on=["pedido", "codigo_equiv"],
    how="outer",
    indicator=True
)

# 🔹 Renombrar para mantener compatibilidad con el resto del código
df_full.rename(columns={"codigo_equiv": "codigo"}, inplace=True)

# ============================================================
# 🔧 Limpieza de duplicados tras merge extendido
# ============================================================
# Eliminar columnas duplicadas (mantiene solo la primera aparición)
df_full = df_full.loc[:, ~df_full.columns.duplicated()].copy()

# En caso de que queden versiones 'codigo_x' o 'codigo_y', unificarlas
if "codigo_x" in df_full.columns:
    df_full["codigo"] = df_full["codigo_x"].combine_first(df_full.get("codigo_y"))
    df_full.drop(columns=["codigo_x", "codigo_y"], errors="ignore", inplace=True)


# # ============================================================
# # 5. CRUCE COMPLETO PARA DETECTAR COINCIDENCIAS Y FALTANTES
# # ============================================================
# df_fenix["origen"] = "FENIX"
# df_elite["origen"] = "ELITE"

# # Cruce completo (outer join)
# df_full = pd.merge(
#     df_fenix,
#     df_elite[["pedido", "codigo", "cantidad_elite", "origen"]],
#     on=["pedido", "codigo"],
#     how="outer",
#     indicator=True
# )

# ============================================================
# 6. GENERAR SUBCONJUNTOS
# ============================================================
# Coincidencias reales (ambos archivos)
df_merge = df_full[df_full["_merge"] == "both"].copy()

# Sin cruce (solo FENIX o solo ELITE)
df_nocruce = df_full[df_full["_merge"] != "both"].copy()
df_nocruce["origen"] = df_nocruce["_merge"].replace({
    "left_only": "Solo en FENIX",
    "right_only": "Solo en ELITE"
})

# ============================================================
# 6.1. REGLA ESPECIAL – Mantener códigos complementarios válidos
# ============================================================
# No enviar a NO_COINCIDEN el código 200492A (ni sus pares)
codigos_validos = ["200492A"]

# Sacar estos registros de df_nocruce y mantenerlos en df_merge
df_extra_validos = df_nocruce[df_nocruce["codigo"].isin(codigos_validos)].copy()
if not df_extra_validos.empty:
    print(f"🧩 Registros especiales mantenidos en CONTROL_ALMACEN: {len(df_extra_validos)}")
    df_extra_validos["estado"] = "OK – Material Complementario"
    df_extra_validos["diferencia"] = 0
    df_merge = pd.concat([df_merge, df_extra_validos], ignore_index=True)

    # Quitar estos del listado de no coincidentes
    df_nocruce = df_nocruce[~df_nocruce["codigo"].isin(codigos_validos)]


# ============================================================
# 7. CÁLCULO DE DIFERENCIA Y ESTADO
# ============================================================
df_merge["cantidad_fenix"] = pd.to_numeric(df_merge.get("cantidad", 0), errors="coerce").fillna(0)
df_merge["cantidad_elite"] = pd.to_numeric(df_merge.get("cantidad_elite", 0), errors="coerce").fillna(0)
df_merge["diferencia"] = df_merge["cantidad_fenix"] - df_merge["cantidad_elite"]

def evaluar(row):
    if row["diferencia"] == 0:
        return "OK"
    elif row["diferencia"] > 0:
        return "FALTANTE EN ELITE"
    else:
        return "EXCESO EN ELITE"

df_merge["estado"] = df_merge.apply(evaluar, axis=1)
# ============================================================
# 7.1. AJUSTE DE MATERIALES COMPLEMENTARIOS (mantiene ambos códigos visibles)
# ============================================================

# 🔹 Diccionario base ↔ complemento (Se puede ampliar sin modificar lógica)
complementos = {
    "200492": "200492A",
    "200384": "200384A"
}

ajustes_realizados = 0

# 🔹 1. Ajuste en df_merge (CONTROL_ALMACEN)
for pedido in df_merge["pedido"].unique():
    for base, comp in complementos.items():
        # Filtrar registros del mismo pedido con el código base o su complemento
        grupo = df_merge[
            (df_merge["pedido"] == pedido)
            & (df_merge["codigo"].isin([base, comp]))
        ]

        if not grupo.empty:
            total_fenix = grupo["cantidad_fenix"].sum()
            total_elite = grupo["cantidad_elite"].sum()

            # Si Elite tiene igual o más cantidad → marcar ambos como complementarios
            if total_elite >= total_fenix and total_fenix > 0:
                df_merge.loc[
                    (df_merge["pedido"] == pedido)
                    & (df_merge["codigo"].isin([base, comp])),
                    ["estado", "diferencia"]
                ] = ["OK – Material Complementario", 0]
                ajustes_realizados += 1

print(f"🔧 Ajustes aplicados (manteniendo ambos códigos): {ajustes_realizados}")

# 🔹 2. Ajuste en df_nocruce (NO_COINCIDEN)
if not df_nocruce.empty:
    registros_ajustados = 0
    for base, comp in complementos.items():
        df_nocruce = df_nocruce[
            ~(
                (df_nocruce["codigo"].isin([base, comp]))
                & (df_nocruce["pedido"].isin(df_merge["pedido"].unique()))
            )
        ]
        registros_ajustados += 1
    print(f"🧩 Registros eliminados de NO_COINCIDEN por complementarios: {registros_ajustados}")
# ============================================================
# 8. ORGANIZAR COLUMNAS FINALES
# ============================================================
columnas_fenix = [
    "pedido", "subz", "municipio", "contrato", "acta",
    "actividad", "fecha_estado", "pagina", "urbrur", "tipre",
    "red_interna", "tipo_operacion", "tipo", "cobro", "suminis",
    "item_cont", "codigo", "cantidad", "vlr_cliente", "valor_costo"
]

# Cambiar el nombre de la columna "estado" a "status" antes del orden
if "estado" in df_merge.columns:
    df_merge.rename(columns={"estado": "status"}, inplace=True)

# De momento NO filtramos columnas aquí — lo haremos al final.
# Esto evita que se pierda la columna 'tecnico' tras el merge.

# ============================================================
# 8.1 AGREGAR COLUMNA TÉCNICO (BUSCARV DESDE PLANILLA CONSUMOS)
# ============================================================
try:
    df_tecnicos = pd.read_excel(ruta_elite, sheet_name=None, dtype=str, header=None)
    hoja_correcta, fila_header = None, None

    for hoja, df_temp in df_tecnicos.items():
        for i, fila in df_temp.iterrows():
            fila_texto = " ".join(str(x).lower() for x in fila.values if pd.notna(x))
            if "tecnico" in fila_texto or "técnico" in fila_texto:
                hoja_correcta, fila_header = hoja, i
                break
        if hoja_correcta:
            break

    if hoja_correcta is None:
        raise Exception("No se encontró ninguna hoja con encabezado 'TECNICO'.")

    df_tecnicos = pd.read_excel(
        ruta_elite,
        sheet_name=hoja_correcta,
        dtype=str,
        skiprows=fila_header
    )

    df_tecnicos.columns = (
        df_tecnicos.columns.map(str)
        .str.lower()
        .str.strip()
        .str.replace(r"unnamed.*", "", regex=True)
    )

    posibles_cols = ["#pedido", "pedido", "codigu", "codigo", "tecnico", "técnico"]
    df_tecnicos = df_tecnicos[[c for c in df_tecnicos.columns if any(p in c for p in posibles_cols)]]

    df_tecnicos.rename(columns={
        "#pedido": "pedido",
        "codigu": "codigo",
        "codigo": "codigo",
        "tecnico": "tecnico",
        "técnico": "tecnico",
    }, inplace=True)

    df_tecnicos = df_tecnicos[["pedido", "tecnico"]].drop_duplicates(subset=["pedido"])

    # 🔹 Merge tipo BUSCARV
    df_merge = df_merge.merge(df_tecnicos, on="pedido", how="left")

    # 🔹 Reemplazar vacíos en la columna técnico por "SIN DATOS"
    if "tecnico" in df_merge.columns:
        df_merge["tecnico"] = df_merge["tecnico"].fillna("SIN DATOS").replace("", "SIN DATOS")

    # 🔹 Reubicar columna 'tecnico' justo después de 'status'
    if "tecnico" in df_merge.columns and "status" in df_merge.columns:
        cols = list(df_merge.columns)
        idx_status = cols.index("status")
        cols.insert(idx_status + 1, cols.pop(cols.index("tecnico")))
        df_merge = df_merge[cols]

    print("👷 Columna 'TÉCNICO' agregada correctamente desde Planilla Consumos.xlsx.")

except Exception as e:
    print(f"⚠️ No se pudo agregar la columna 'TÉCNICO': {e}")

# ============================================================
# 8.2 ORDEN FINAL DE COLUMNAS (ya con TÉCNICO incluido)
# ============================================================
columnas_finales = columnas_fenix + ["cantidad_elite", "diferencia", "status", "tecnico"]
df_merge = df_merge[[c for c in columnas_finales if c in df_merge.columns]]

# Para hoja NO_COINCIDEN
columnas_nocruce = ["pedido", "codigo", "cantidad", "cantidad_elite", "origen"]
df_nocruce = df_nocruce[[c for c in columnas_nocruce if c in df_nocruce.columns]]
# ============================================================
# 8.3 RECONSTRUCCIÓN FINAL DE HOJA NO_COINCIDEN (v4.0 con cantidad real)
# ============================================================
try:
    # --- Leer planilla para obtener pedido, código, cantidad y técnico ---
    df_planilla = pd.read_excel(ruta_elite, sheet_name=None, dtype=str, header=None)
    hoja_correcta, fila_header = None, None

    for hoja, df_temp in df_planilla.items():
        for i, fila in df_temp.iterrows():
            texto = " ".join(str(x).lower() for x in fila.values if pd.notna(x))
            if "tecnico" in texto or "técnico" in texto:
                hoja_correcta, fila_header = hoja, i
                break
        if hoja_correcta:
            break

    if not hoja_correcta:
        raise Exception("No se encontró hoja con columna técnico.")

    # Leer desde encabezado detectado
    df_planilla = pd.read_excel(
        ruta_elite,
        sheet_name=hoja_correcta,
        dtype=str,
        skiprows=fila_header
    )

    # Normalizar encabezados
    df_planilla.columns = (
        df_planilla.columns.map(str)
        .str.lower()
        .str.strip()
        .str.replace(r"unnamed.*", "", regex=True)
    )

    # Renombrar columnas clave
    df_planilla.rename(columns={
        "#pedido": "pedido",
        "codigu": "codigo",
        "cantidad": "cantidad_elite",
        "técnico": "tecnico"
    }, inplace=True)

    # Filtrar columnas relevantes
    columnas_necesarias = ["pedido", "codigo", "cantidad_elite", "tecnico"]
    df_planilla = df_planilla[[c for c in df_planilla.columns if c in columnas_necesarias]].copy()

    # Limpieza básica
    df_planilla["pedido"] = df_planilla["pedido"].astype(str).str.strip()
    df_planilla["codigo"] = df_planilla["codigo"].astype(str).str.strip()
    df_planilla["tecnico"] = df_planilla["tecnico"].astype(str).str.strip()
    df_planilla["cantidad_elite"] = (
    df_planilla["cantidad_elite"]
    .astype(str)
    .str.replace(",", ".", regex=False)
    .apply(lambda x: float(x) if x.replace(".", "", 1).isdigit() else 0)
)

    df_planilla.dropna(subset=["pedido", "codigo"], inplace=True)
    df_planilla.drop_duplicates(subset=["pedido", "codigo"], keep="first", inplace=True)

    # --- Filtrar registros Solo en ELITE ---
    df_nc_elite = df_nocruce[df_nocruce["origen"].str.contains("Solo en ELITE", case=False, na=False)].copy()
    df_nc_otros = df_nocruce[~df_nocruce["origen"].str.contains("Solo en ELITE", case=False, na=False)].copy()

    if not df_nc_elite.empty:
        pedidos_elite = df_nc_elite["pedido"].unique().tolist()
        df_codigos_planilla = df_planilla[df_planilla["pedido"].isin(pedidos_elite)].copy()

        # Crear base limpia con estructura correcta
        df_nueva_elite = pd.DataFrame({
            "pedido": df_codigos_planilla["pedido"],
            "codigo": df_codigos_planilla["codigo"],
            "cantidad": 0,
            "cantidad_elite": df_codigos_planilla["cantidad_elite"],
            "origen": "Solo en ELITE",
            "tecnico": df_codigos_planilla["tecnico"]
        })

        # Evitar duplicados reales
        df_nueva_elite.drop_duplicates(subset=["pedido", "codigo"], keep="first", inplace=True)

        # Combinar con el resto (Solo FENIX, etc.)
        df_nocruce = pd.concat([df_nc_otros, df_nueva_elite], ignore_index=True)

    # 🔹 Asegurar orden de columnas
    columnas_nocruce = ["pedido", "codigo", "cantidad", "cantidad_elite", "origen", "tecnico"]
    df_nocruce = df_nocruce[[c for c in columnas_nocruce if c in df_nocruce.columns]]

    print("✅ Hoja NO_COINCIDEN reconstruida con cantidades reales y técnico correcto (v4.0).")

except Exception as e:
    print(f"⚠️ Error al reconstruir hoja NO_COINCIDEN: {e}")

# ============================================================
# 9. CREAR RESUMEN
# ============================================================
# Agrupamos por la nueva columna "status" en lugar de "estado_final"
resumen = (
    df_merge.groupby("status", dropna=False)
    .size()
    .reset_index(name="total")
    .sort_values(by="status")
)
total_registros = len(df_merge)
resumen.loc[len(resumen)] = ["TOTAL GENERAL", total_registros]

# Cambiamos nombre de la columna del resumen a "estado_final"
resumen.rename(columns={"status": "estado_final"}, inplace=True)

# ============================================================
# 🔹 LIMPIEZA DE PEDIDOS (evita falsos pedidos 1, 2, 3…)
# ============================================================
if "pedido" in df_elite.columns:
    # Normalizar y eliminar filas sin pedido válido
    df_elite["pedido"] = (
        df_elite["pedido"]
        .astype(str)
        .str.strip()
        .replace({"nan": None, "": None})
    )

    # Conservar solo filas con pedidos numéricos reales de 8 dígitos o más
    df_elite = df_elite[
        df_elite["pedido"].notna() &
        df_elite["pedido"].str.match(r"^\d{8,}$", na=False)
    ]

    # Eliminar filas vacías restantes
    df_elite = df_elite.dropna(subset=["pedido"])

# ============================================================
# 🔧 Limpieza final: evitar pedidos duplicados entre FÉNIX y ELITE
# ============================================================
try:
    if 'df_nocruce' in locals() and not df_nocruce.empty:
        # Asegurar tipos de datos consistentes
        df_nocruce["pedido"] = df_nocruce["pedido"].astype(str).str.strip()
        df_nocruce["origen"] = df_nocruce["origen"].astype(str)

        # 1️⃣ Obtener todos los pedidos que están en "Solo en ELITE"
        pedidos_elite = df_nocruce.loc[
            df_nocruce["origen"].str.contains("Solo en ELITE", case=False, na=False),
            "pedido"
        ].unique()

        # 2️⃣ Eliminar versiones duplicadas de esos mismos pedidos en "Solo en FENIX"
        df_nocruce = df_nocruce[
            ~(
                (df_nocruce["pedido"].isin(pedidos_elite)) &
                (df_nocruce["origen"].str.contains("Solo en FENIX", case=False, na=False))
            )
        ].copy()

        # 3️⃣ Eliminar duplicados exactos (por pedido + código)
        df_nocruce.drop_duplicates(subset=["pedido", "codigo"], keep="first", inplace=True)

        # 4️⃣ Ordenar por pedido y código
        df_nocruce.sort_values(by=["pedido", "codigo"], inplace=True, ignore_index=True)

        print("🧩 Limpieza aplicada: eliminados duplicados FÉNIX/ELITE por pedido (v4.4).")
    else:
        print("⚠️ df_nocruce vacío o no definido, se omite limpieza final.")
except Exception as e:
    print(f"⚠️ Error al limpiar duplicados entre FÉNIX y ELITE: {e}")

# ============================================================
# 10. EXPORTAR A EXCEL (manejo de archivo abierto)
# ============================================================


try:
    ruta_salida.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(ruta_salida, engine="openpyxl") as writer:
        df_merge.to_excel(writer, index=False, sheet_name="CONTROL_ALMACEN")
        resumen.to_excel(writer, index=False, sheet_name="RESUMEN")
        df_nocruce.to_excel(writer, index=False, sheet_name="NO_COINCIDEN")

    print("💾 Exportando archivo con hoja de control de pendientes...")

except PermissionError:
    print("⚠️ No se puede guardar el archivo porque está abierto en Excel.")
    print("🧩 Por favor, cierre 'CONTROL_ALMACEN.xlsx' y ejecute nuevamente el script.")
    import sys
    sys.exit(1)

except Exception as e:
    print(f"❌ Error inesperado al exportar a Excel: {e}")
    import sys
    sys.exit(1)


# ============================================================
# 🔹 NORMALIZAR TIPOS DE DATOS (evita "Recuento" en Excel)
# ============================================================
cols_numericas = ["cantidad", "cantidad_elite", "vlr_cliente", "valor_costo", "diferencia"]

for col in cols_numericas:
    if col in df_merge.columns:
        df_merge[col] = (
            pd.to_numeric(df_merge[col], errors="coerce")
            .fillna(0)
            .astype(float)
        )
# ============================================================
# 🔹 Normalizar tipos numéricos también en NO_COINCIDEN
# ============================================================
cols_numericas_nc = ["cantidad", "cantidad_elite"]

for col in cols_numericas_nc:
    if col in df_nocruce.columns:
        df_nocruce[col] = (
            pd.to_numeric(df_nocruce[col], errors="coerce")
            .fillna(0)
            .astype(float)
        )

with pd.ExcelWriter(ruta_salida, engine="openpyxl") as writer:
    df_merge.to_excel(writer, index=False, sheet_name="CONTROL_ALMACEN")
    resumen.to_excel(writer, index=False, sheet_name="RESUMEN")
    df_nocruce.to_excel(writer, index=False, sheet_name="NO_COINCIDEN")

print("💾 Exportando archivo con hoja de control de pendientes...")

# Asegurar que las columnas numéricas estén en formato numérico real
for col in ["cantidad", "cantidad_elite", "vlr_cliente", "valor_costo", "diferencia"]:
    if col in df_merge.columns:
        df_merge[col] = pd.to_numeric(df_merge[col], errors="coerce").fillna(0)

# ============================================================
# 11. FORMATO VISUAL LIMPIO
# ============================================================
wb = load_workbook(ruta_salida)

def formato_hoja(ws):
    from openpyxl.styles import PatternFill, Font, Alignment

    max_row, max_col = ws.max_row, ws.max_column

    font_encabezado = Font(color="FFFFFF", bold=True)
    align_center = Alignment(horizontal="center", vertical="center")

    # 🎨 Paleta de colores
    colores = {
        "default": "004C99",      # azul (FENIX)
        "elite": "000000",        # negro (ELITE)
        "diferencia": "000000",   # negro (comparativo)
        "status": "000000",       # negro (resultado)
        "tecnico": "000000",      # negro (nueva columna técnico)
    }

    # 🔹 Colorear encabezados según tipo
    for idx, cell in enumerate(ws[1], 1):
        header = str(cell.value).lower().strip()
        color = colores["default"]  # por defecto azul FENIX

        if "elite" in header:
            color = colores["elite"]
        elif "diferencia" in header:
            color = colores["diferencia"]
        elif header == "status":  # evitar confusión con fecha_estado
            color = colores["status"]
        elif "tecnico" in header:
            color = colores["tecnico"]    

        cell.fill = PatternFill("solid", start_color=color)
        cell.font = font_encabezado
        cell.alignment = align_center

    # 🔹 Alinear celdas del cuerpo
    for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=max_col):
        for c in row:
            c.alignment = align_center


# === CONTROL_ALMACEN ===
ws = wb["CONTROL_ALMACEN"]
formato_hoja(ws)

# Aplicar semáforo sobre columna STATUS
col_status = None
for idx, cell in enumerate(ws[1], 1):
    if str(cell.value).lower().strip() == "status":
        col_status = idx
        break

if col_status:
    for i in range(2, ws.max_row + 1):
        c = ws.cell(row=i, column=col_status)
        text = str(c.value).upper()
        if "OK" in text:
            c.fill = PatternFill("solid", start_color="00B050")
            c.font = Font(color="FFFFFF", bold=True)
        elif "FALTANTE" in text:
            c.fill = PatternFill("solid", start_color="FFD966")
            c.font = Font(color="000000", bold=True)
        elif "EXCESO" in text:
            c.fill = PatternFill("solid", start_color="C00000")
            c.font = Font(color="FFFFFF", bold=True)

# === RESUMEN ===
ws_resumen = wb["RESUMEN"]
formato_hoja(ws_resumen)

# === NO_COINCIDEN ===
ws_nc = wb["NO_COINCIDEN"]
formato_hoja(ws_nc)

for i in range(2, ws_nc.max_row + 1):
    c = ws_nc.cell(row=i, column=ws_nc.max_column)
    if "ELITE" in str(c.value).upper():
        c.fill = PatternFill("solid", start_color="C00000")
        c.font = Font(color="FFFFFF", bold=True)
    elif "FENIX" in str(c.value).upper():
        c.fill = PatternFill("solid", start_color="1F4E78")
        c.font = Font(color="FFFFFF", bold=True)

wb.save(ruta_salida)
wb.close()

print("✅ CRUCE FINALIZADO CON ÉXITO (v3.7 con colores de encabezado).")
print(f"📁 Archivo generado: {ruta_salida}")
print("------------------------------------------------------------")

	